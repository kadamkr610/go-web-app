import hvac
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

import hvac.exceptions

VAULT_ADDR = "https://hcp-vault-private-vault-fc507e0d.5d5b1f21.z1.hashicorp.cloud:8200"
VAULT_TOKEN = os.getenv("VAULT_TOKEN")
TTL_THRESHOLD = 31536000  # 1 year in seconds

client = hvac.Client(url=VAULT_ADDR, token=VAULT_TOKEN)
print(client.is_authenticated())

def list_namespaces(namespace=""):
    client.adapter.namespace = namespace
    try:
        all_namespaces = []
        response = client.sys.list_namespaces()
        namespaces = response.get("data", {}).get("keys", [])
        print(namespaces)
        full_namespaces = [namespace + "/" + ns.strip("/") for ns in namespaces] # Ensure no leading/trailing slashes
        for ns in full_namespaces:
            all_namespaces.append(ns)
            all_namespaces.extend(list_namespaces(ns))
        return all_namespaces
    except hvac.exceptions.InvalidPath: #If invalid, it means we are at the root or no namespaces exist
        return all_namespaces
    except hvac.exceptions.Forbidden:
        return all_namespaces

def check_approle_secret_id(namespace, rownum):
    workbook = openpyxl.load_workbook("./auth_secret_id_TTL.xlsx")

    sheet = workbook.active = workbook["Approle Secret ID"] #Change to Approle Secret ID sheet
    client.adapter.namespace = namespace #Change to namespace from parameter
    current_row = rownum
    rows = []
    try:
        roles = client.auth.approle.list_roles()["data"]["keys"]
        print(roles)

        for role in roles: #Iterate through each role and if secret_id_ttl is greater than threshold
            print(role)
            try:
                secret_id_accessors = client.auth.approle.list_secret_id_accessors(role_name=role)["data"]["keys"]
                for accessor in secret_id_accessors:
                    secret_id_info = client.auth.approle.read_secret_id_accessor(role_name=role, secret_id_accessor=accessor)
                    ttl = secret_id_info["data"]["secret_id_ttl"]
                    if ttl == 0 or ttl > TTL_THRESHOLD:
                        rows.append([namespace, role, accessor, ttl])
                        print(f"[!] Namespace: '{namespace or 'root'}', Role: '{role}', Secret ID Accessor: {accessor}, TTL: {ttl} seconds")
            except hvac.exceptions.InvalidPath:
                pass
    except hvac.exceptions.InvalidPath:
        pass
    except Exception as e:
        print(f"Error in namespace '{namespace}': {e}")

    for row_index, row_data in enumerate(rows, current_row): #Input values into sheet
        current_row+=1
        for col_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    for column_cells in sheet.columns:
      new_column_length = max(len(str(cell.value)) for cell in column_cells)
      new_column_letter = (chr(64+(column_cells[0].column)))
      if new_column_length > 0:
        sheet.column_dimensions[new_column_letter].width = new_column_length # *1.10

    workbook.save(filename="./auth_secret_id_TTL.xlsx")
    return current_row
    
def check_approle_roles(namespace, rownum):
    workbook = openpyxl.load_workbook("./auth_secret_id_TTL.xlsx")

    sheet = workbook.active = workbook["Approle Roles"]
    client.adapter.namespace = namespace #Change to namespace from parameter
    current_row = rownum
    rows = []
    try:
        roles = client.auth.approle.list_roles()["data"]["keys"]

        for role in roles: #Go through each role in approle and check the secret_id_ttl settings
            try:
                role_id_data = client.auth.approle.read_role(role_name=role)["data"] #Grab role metadata
                ttl = role_id_data["secret_id_ttl"]
                if ttl == 0 or ttl > TTL_THRESHOLD: #If ttl is 0 or greater than threshold, add to spreadsheet
                    rows.append([namespace, role, ttl])
                    print(f"[!] Namespace: '{namespace or 'root'}', Role: '{role}',TTL: {ttl} seconds")
            except hvac.exceptions.InvalidPath: #Skip invalid path as it means role can potentially not exist or have metadata
                pass
    except hvac.exceptions.InvalidPath:
        pass
    except Exception as e:
        print(f"Error in namespace '{namespace}': {e}")

    for row_index, row_data in enumerate(rows, current_row): #Input values into sheet
        current_row+=1
        for col_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    for column_cells in sheet.columns: #Adjust column widths based on content
      new_column_length = max(len(str(cell.value)) for cell in column_cells)
      new_column_letter = (chr(64+(column_cells[0].column)))
      if new_column_length > 0:
        sheet.column_dimensions[new_column_letter].width = new_column_length # *1.10

    workbook.save(filename="./auth_secret_id_TTL.xlsx")
    return current_row #Return the next row number to write to for the next function call

def list_of_auth_methods(namespace, rownum):
    client.adapter.namespace = namespace
    workbook = openpyxl.load_workbook("./auth_secret_id_TTL.xlsx")

    sheet = workbook.active = workbook["Authentication Methods"]
    approle_fill = PatternFill(start_color="FEDCAA", end_color="FEDCAA", fill_type="solid") #Peach color for highling AppRole auth methods
    rows = []
    current_row = rownum #Keep track of the current row for writing data

    print(client.sys.list_auth_methods())

    auth_methods = client.sys.list_auth_methods()

    for path, data in auth_methods["data"].items(): #Grab all auth method paths and their types
        auth_type = data.get('type')
        rows.append([namespace, path, auth_type])
        print(f"- Path: {path}")
        print(f"  Type: {auth_type}")
        print(rows)
    
    for row_index, row_data in enumerate(rows, current_row): #Add the auth methods to the sheet
        current_row+=1
        for col_index, value in enumerate(row_data, start=1):
            if value == "approle":
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.fill = approle_fill
            else:
                sheet.cell(row=row_index, column=col_index, value=value)
    
    for column_cells in sheet.columns: #Adjust column widths based on content
      new_column_length = max(len(str(cell.value)) for cell in column_cells)
      new_column_letter = (chr(64+(column_cells[0].column)))
      if new_column_length > 0:
        sheet.column_dimensions[new_column_letter].width = new_column_length # *1.10

    workbook.save(filename="auth_secret_id_TTL.xlsx")
    return current_row



def main():
    if not client.is_authenticated():
        print("Vault authentication failed.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title="Approle Secret ID"
    approle_headers = ["NameSpace", "Role", "Accessor", "TTL"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

    for col_index, header in enumerate(approle_headers, start=1): #Create headers for Approle Secret ID sheet
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill

    
    auth_headers = ["Namespace", "Mount Path", "Type"]
    workbook.create_sheet("Authentication Methods")
    sheet = workbook.active = workbook["Authentication Methods"]
    for col_index, header in enumerate(auth_headers, start=1): #Create headers for Authentication Methods sheet
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
    workbook.save(filename="./auth_secret_id_TTL.xlsx")

    auth_headers = ["Namespace", "Role", "TTL"]
    workbook.create_sheet("Approle Roles")
    sheet = workbook.active = workbook["Approle Roles"]
    for col_index, header in enumerate(auth_headers, start=1): #Create headers for Approle Roles sheet
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
    workbook.save(filename="./auth_secret_id_TTL.xlsx")
    print("Scanning Vault namespaces for AppRole roles with long-lived secret_id_ttl...\n")
    namespaces = list_namespaces("admin")
    print(namespaces)
    rownum_secret_id = 2 # Start writing from the second row due to headers being in the first row
    rownum_approles = 2
    rownum_auth = 2

    for ns in namespaces:
        rownum_secret_id = check_approle_secret_id(ns, rownum_secret_id)
        rownum_approles = check_approle_roles(ns, rownum_approles)
        rownum_auth = list_of_auth_methods(ns, rownum_auth)

if __name__ == "__main__":
    main()
